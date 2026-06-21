import os
import asyncio
import openpyxl
from typing import Optional, List, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from .. import models
from ..logger import logger

# Список известных кодов языков и их синонимов для импорта из Excel
LANG_ALIASES = {
    'word': 'en', 'eng': 'en', 'english': 'en',
    'de': 'de', 'german': 'de', 'deutsch': 'de',
    'it': 'it', 'italian': 'it', 'italiano': 'it',
    'ru': 'ru', 'russian': 'ru', 'русский': 'ru',
    'meaning': 'meaning', 'context': 'meaning', 'значение': 'meaning'
}

def clean_excel_val(val: Any) -> str:
    """Очищает значение ячейки Excel от лишних пробелов и None."""
    if val is None:
        return ""
    return str(val).strip()

class SyncService:
    """Сервис для синхронизации данных между БД и внешними файлами (Excel)."""
    
    def __init__(self, db: AsyncSession):
        self.db = db

    async def import_excel_to_db(self, excel_path: str) -> Dict[str, Any]:
        """Импортирует слова из Excel в таблицу WordStats."""
        db = self.db
        if not os.path.exists(excel_path):
            return {
                "success": False,
                "message": f"Excel file not found at path: {excel_path}",
                "counts": {"new": 0, "updated": 0, "skipped": 0, "conflicts": 0},
                "conflicts_list": []
            }

        try:
            loop = asyncio.get_running_loop()

            def read_excel():
                wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                sheet = wb.active
                if sheet is None:
                    return []
                rows = list(sheet.iter_rows(values_only=True))
                wb.close()
                return rows

            rows = await loop.run_in_executor(None, read_excel)

            if not rows:
                return {
                    "success": False,
                    "message": "File is empty",
                    "counts": {"new": 0, "updated": 0, "skipped": 0, "conflicts": 0},
                    "conflicts_list": []
                }
            
            counts = {"new": 0, "updated": 0, "skipped": 0, "conflicts": 0}
            conflicts_list: List[Dict[str, Any]] = []

            # Определяем маппинг колонок по заголовку
            header = rows[0]
            col_map: Dict[int, str] = {} # index -> lang_code
            
            has_header = False
            for i, cell in enumerate(header):
                val = clean_excel_val(cell).lower()
                if val in LANG_ALIASES:
                    col_map[i] = LANG_ALIASES[val]
                    has_header = True
                elif len(val) == 2: # Предполагаем, что 2 буквы - это код языка
                    col_map[i] = val
                    has_header = True

            # Если заголовок не распознан, используем дефолтный маппинг
            if not has_header:
                col_map = {0: 'en', 1: 'de', 2: 'it', 3: 'ru', 4: 'meaning'}
                data_rows = rows
            else:
                data_rows = rows[1:]

            for r in data_rows:
                if not r: continue
                
                # Извлекаем данные по маппингу
                row_data: Dict[str, str] = {}
                for i, lang_code in col_map.items():
                    if i < len(r):
                        row_data[lang_code] = clean_excel_val(r[i])

                eng = row_data.get('en', '')
                if not eng or eng.lower() in ['английский', 'english', 'word', 'eng']:
                    continue

                meaning = row_data.pop('meaning', '')
                ru = row_data.get('ru', '')
                it = row_data.get('it', '')
                de = row_data.get('de', '')

                # Ищем существующее слово
                existing_res = await db.execute(select(models.WordStats).where(models.WordStats.word == eng))
                existing = existing_res.scalar_one_or_none()

                if existing:
                    # Сравниваем переводы
                    db_trans = existing.translations or {}
                    db_meaning = existing.meaning or ""
                    
                    field_diffs: Dict[str, Dict[str, str]] = {}
                    for lang, val in row_data.items():
                        db_val = db_trans.get(lang, "")
                        if db_val != val:
                            field_diffs[lang] = {'db': db_val, 'file': val}
                    
                    if db_meaning != meaning:
                        field_diffs['meaning'] = {'db': db_meaning, 'file': meaning}

                    if field_diffs:
                        counts["conflicts"] += 1
                        conflicts_list.append({
                            "word": eng,
                            "field_diffs": field_diffs
                        })
                    else:
                        counts["skipped"] += 1
                else:
                    # Новое слово
                    db.add(models.WordStats(
                        word=eng, eng=eng, it=it, de=de, ru=ru, 
                        meaning=meaning, translations=row_data,
                        knowledge_stats={l: False for l in row_data.keys()}
                    ))
                    counts["new"] += 1

            await db.commit()
            logger.info(f"Import summary: {counts}")

            message = f"Import completed. New: {counts['new']}, Skipped: {counts['skipped']}, Conflicts: {counts['conflicts']}"
            return {
                "success": True,
                "message": message,
                "counts": counts,
                "conflicts_list": conflicts_list
            }
        except Exception as e:
            await db.rollback()
            logger.error(f"Excel import error: {e}", exc_info=True)
            return {
                "success": False,
                "message": f"Error during import: {str(e)}",
                "counts": {"new": 0, "updated": 0, "skipped": 0, "conflicts": 0},
                "conflicts_list": []
            }

    async def sync_conflicts(self, excel_path: str, resolutions: Dict[str, str]) -> Dict[str, Any]:
        """Применяет разрешения конфликтов."""
        db = self.db
        if not os.path.exists(excel_path):
            return {"success": False, "message": "Excel file not found."}

        try:
            words_to_sync = list(resolutions.keys())
            if not words_to_sync:
                return {"success": True, "message": "No data for synchronization.", "updated_db": 0, "updated_file": 0}

            db_records_res = await db.execute(select(models.WordStats).where(models.WordStats.word.in_(words_to_sync)))
            db_records = db_records_res.scalars().all()
            db_map = {r.word: r for r in db_records}

            loop = asyncio.get_running_loop()

            def process_excel():
                wb = openpyxl.load_workbook(excel_path)
                sheet = wb.active
                if sheet is None: return 0, {}
                rows = list(sheet.iter_rows())
                if not rows: return 0, {}

                header = [clean_excel_val(c.value).lower() for c in rows[0]]
                col_map: Dict[int, str] = {}
                for i, val in enumerate(header):
                    if val in LANG_ALIASES: col_map[i] = LANG_ALIASES[val]
                    elif len(val) == 2: col_map[i] = val

                updated_file_count = 0
                db_updates: Dict[str, Dict[str, Any]] = {}

                for row_idx, row in enumerate(rows[1:], start=2):
                    if not row or row[0].value is None: continue
                    eng = clean_excel_val(row[0].value)
                    if not eng or eng not in resolutions: continue

                    action = resolutions[eng]
                    if action == 'to_file' and eng in db_map:
                        rec = db_map[eng]
                        trans = rec.translations or {}
                        for col_idx, lang_code in col_map.items():
                            if lang_code == 'en': continue
                            if lang_code == 'meaning':
                                sheet.cell(row=row_idx, column=col_idx+1, value=rec.meaning)
                            else:
                                sheet.cell(row=row_idx, column=col_idx+1, value=trans.get(lang_code, ""))
                        updated_file_count += 1
                    
                    elif action == 'to_db':
                        new_trans: Dict[str, str] = {}
                        new_meaning = ""
                        for col_idx, lang_code in col_map.items():
                            val = clean_excel_val(row[col_idx].value)
                            if lang_code == 'meaning': new_meaning = val
                            else: new_trans[lang_code] = val
                        db_updates[eng] = {'translations': new_trans, 'meaning': new_meaning}

                if updated_file_count > 0:
                    wb.save(excel_path)
                wb.close()
                return updated_file_count, db_updates

            file_updates_count, db_updates_data = await loop.run_in_executor(None, process_excel)

            db_updates_count = 0
            for eng, data in db_updates_data.items():
                if eng in db_map:
                    rec = db_map[eng]
                    rec.meaning = data['meaning']
                    current_trans = dict(rec.translations or {})
                    current_trans.update(data['translations'])
                    rec.translations = current_trans
                    
                    # Back-compat columns
                    rec.ru = data['translations'].get('ru', rec.ru)
                    rec.it = data['translations'].get('it', rec.it)
                    rec.de = data['translations'].get('de', rec.de)
                    db_updates_count += 1

            if db_updates_count > 0:
                await db.commit()

            return {
                "success": True,
                "message": f"Synchronization completed. DB: {db_updates_count}, File: {file_updates_count}.",
                "updated_db": db_updates_count,
                "updated_file": file_updates_count
            }

        except PermissionError:
            return {"success": False, "message": "Access error: close the Excel file."}
        except Exception as e:
            await db.rollback()
            logger.error(f"Sync conflict error: {e}", exc_info=True)
            return {"success": False, "message": f"Synchronization error: {str(e)}"}
