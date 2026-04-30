import os
import json
import asyncio
import openpyxl
from datetime import datetime, date, timezone
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, text
from typing import Optional, List, Dict, Any, Tuple, Union

from .. import models
from .settings_service import get_setting, set_settings_batch
from ..logger import logger

_state_locks: Dict[asyncio.AbstractEventLoop, asyncio.Lock] = {}

def _get_state_lock() -> asyncio.Lock:
    """Лази-инициализация лока, привязанного к текущему циклу событий."""
    loop = asyncio.get_running_loop()
    if loop not in _state_locks:
        _state_locks[loop] = asyncio.Lock()
    return _state_locks[loop]

def clean_excel_val(val: Any) -> str:
    """Очищает значение ячейки Excel от лишних пробелов и None."""
    if val is None:
        return ""
    return str(val).strip()

class StateManager:
    """Сервис для управления глобальным состоянием и импортом данных."""
    
    def __init__(self, db: AsyncSession):
        self.db = db

    async def import_excel_to_db(self, excel_path: str) -> Dict[str, Any]:
        """Импортирует слова из Excel в таблицу WordStats."""
        db = self.db
        if not os.path.exists(excel_path):
            return {
                "success": False,
                "message": f"Файл Excel не найден по пути: {excel_path}",
                "counts": {"new": 0, "updated": 0, "skipped": 0, "conflicts": 0},
                "conflicts_list": []
            }

        try:
            loop = asyncio.get_event_loop()

            def read_excel():
                wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                sheet = wb.active
                if sheet is None:
                    return []
                rows = list(sheet.iter_rows(values_only=True))
                wb.close()
                return rows

            rows = await loop.run_in_executor(None, read_excel)

            if not rows:
                return {
                    "success": False,
                    "message": "Файл пуст",
                    "counts": {"new": 0, "updated": 0, "skipped": 0, "conflicts": 0},
                    "conflicts_list": []
                }
            
            counts = {"new": 0, "updated": 0, "skipped": 0, "conflicts": 0}
            conflicts_list: List[Dict[str, Any]] = []

            # Определяем маппинг колонок по заголовку
            header = rows[0]
            col_map: Dict[int, str] = {} # index -> lang_code
            
            # Список известных кодов и их синонимов
            lang_aliases = {
                'word': 'en', 'eng': 'en', 'english': 'en',
                'de': 'de', 'german': 'de', 'deutsch': 'de',
                'it': 'it', 'italian': 'it', 'italiano': 'it',
                'ru': 'ru', 'russian': 'ru', 'русский': 'ru',
                'meaning': 'meaning', 'context': 'meaning', 'значение': 'meaning'
            }

            has_header = False
            for i, cell in enumerate(header):
                val = clean_excel_val(cell).lower()
                if val in lang_aliases:
                    col_map[i] = lang_aliases[val]
                    has_header = True
                elif len(val) == 2: # Предполагаем, что 2 буквы - это код языка (fr, es, la и т.д.)
                    col_map[i] = val
                    has_header = True

            # Если заголовок не распознан, используем дефолтный маппинг
            if not has_header:
                col_map = {0: 'en', 1: 'de', 2: 'it', 3: 'ru', 4: 'meaning'}
                data_rows = rows
            else:
                data_rows = rows[1:]

            for r in data_rows:
                if not r: continue
                
                # Извлекаем данные по маппингу
                row_data: Dict[str, str] = {}
                for i, lang_code in col_map.items():
                    if i < len(r):
                        row_data[lang_code] = clean_excel_val(r[i])

                eng = row_data.get('en', '')
                if not eng or eng.lower() in ['английский', 'english', 'word', 'eng']:
                    continue

                meaning = row_data.pop('meaning', '')
                ru = row_data.get('ru', '') # Сохраняем для обратной совместимости колонок
                it = row_data.get('it', '')
                de = row_data.get('de', '')

                # Ищем существующее слово
                existing_res = await db.execute(select(models.WordStats).where(models.WordStats.word == eng))
                existing = existing_res.scalar_one_or_none()

                if existing:
                    # Сравниваем переводы
                    db_trans = existing.translations or {}
                    db_meaning = existing.meaning or ""
                    
                    field_diffs: Dict[str, Dict[str, str]] = {}
                    # Проверяем все языки из файла
                    for lang, val in row_data.items():
                        db_val = db_trans.get(lang, "")
                        if db_val != val:
                            field_diffs[lang] = {'db': db_val, 'file': val}
                    
                    # Проверяем описание
                    if db_meaning != meaning:
                        field_diffs['meaning'] = {'db': db_meaning, 'file': meaning}

                    if field_diffs:
                        counts["conflicts"] += 1
                        conflicts_list.append({
                            "word": eng,
                            "field_diffs": field_diffs
                        })
                    else:
                        counts["skipped"] += 1
                else:
                    # Новое слово
                    db.add(models.WordStats(
                        word=eng, eng=eng, it=it, de=de, ru=ru, 
                        meaning=meaning, translations=row_data,
                        knowledge_stats={l: False for l in row_data.keys()}
                    ))
                    counts["new"] += 1

            await db.commit()
            logger.info(f"Import summary: {counts}")

            message = f"Импорт завершен. Новых: {counts['new']}, Пропущено (совпадает): {counts['skipped']}, Конфликтов: {counts['conflicts']}"
            if counts["conflicts"] > 0:
                message += ". Проверьте конфликты перед повторным импортом."

            return {
                "success": True,
                "message": message,
                "counts": counts,
                "conflicts_list": conflicts_list
            }
        except Exception as e:
            await db.rollback()
            logger.error(f"Excel import error: {e}", exc_info=True)
            return {
                "success": False,
                "message": f"Ошибка при импорте: {str(e)}",
                "counts": {"new": 0, "updated": 0, "skipped": 0, "conflicts": 0},
                "conflicts_list": []
            }

    async def sync_conflicts(self, excel_path: str, resolutions: Dict[str, str]) -> Dict[str, Any]:
        """Применяет разрешения конфликтов: записывает данные в БД или изменяет файл Excel."""
        db = self.db
        if not os.path.exists(excel_path):
            return {"success": False, "message": "Файл Excel не найден."}

        try:
            words_to_sync = list(resolutions.keys())
            if not words_to_sync:
                return {"success": True, "message": "Нет данных для синхронизации.", "updated_db": 0, "updated_file": 0}

            db_records_res = await db.execute(select(models.WordStats).where(models.WordStats.word.in_(words_to_sync)))
            db_records = db_records_res.scalars().all()
            
            # Подготавливаем словарь для быстрого поиска
            db_map = {r.word: r for r in db_records}

            loop = asyncio.get_event_loop()

            def process_excel():
                wb = openpyxl.load_workbook(excel_path)
                sheet = wb.active
                if sheet is None: return 0, {}
                rows = list(sheet.iter_rows())
                if not rows: return 0, {}

                # Определяем маппинг колонок по заголовку
                header = [clean_excel_val(c.value).lower() for c in rows[0]]
                lang_aliases = {
                    'word': 'en', 'eng': 'en', 'english': 'en',
                    'de': 'de', 'german': 'de', 'deutsch': 'de',
                    'it': 'it', 'italian': 'it', 'italiano': 'it',
                    'ru': 'ru', 'russian': 'ru', 'русский': 'ru',
                    'meaning': 'meaning', 'context': 'meaning', 'значение': 'meaning'
                }
                
                col_map: Dict[int, str] = {}
                for i, val in enumerate(header):
                    if val in lang_aliases: col_map[i] = lang_aliases[val]
                    elif len(val) == 2: col_map[i] = val

                updated_file_count = 0
                db_updates: Dict[str, Dict[str, Any]] = {} # word -> {translations: {}, meaning: ""}

                for row_idx, row in enumerate(rows[1:], start=2):
                    if not row or row[0].value is None: continue
                    eng = clean_excel_val(row[0].value)
                    if not eng or eng not in resolutions: continue

                    action = resolutions[eng]
                    if action == 'to_file' and eng in db_map:
                        rec = db_map[eng]
                        trans = rec.translations or {}
                        # Обновляем все колонки в файле, которые есть в маппинге
                        for col_idx, lang_code in col_map.items():
                            if lang_code == 'en': continue
                            if lang_code == 'meaning':
                                sheet.cell(row=row_idx, column=col_idx+1, value=rec.meaning)
                            else:
                                sheet.cell(row=row_idx, column=col_idx+1, value=trans.get(lang_code, ""))
                        updated_file_count += 1
                    
                    elif action == 'to_db':
                        new_trans: Dict[str, str] = {}
                        new_meaning = ""
                        for col_idx, lang_code in col_map.items():
                            val = clean_excel_val(row[col_idx].value)
                            if lang_code == 'meaning': new_meaning = val
                            else: new_trans[lang_code] = val
                        db_updates[eng] = {'translations': new_trans, 'meaning': new_meaning}

                if updated_file_count > 0:
                    wb.save(excel_path)
                wb.close()
                return updated_file_count, db_updates

            file_updates_count, db_updates_data = await loop.run_in_executor(None, process_excel)

            db_updates_count = 0
            for eng, data in db_updates_data.items():
                if eng in db_map:
                    rec = db_map[eng]
                    rec.meaning = data['meaning']
                    
                    # Мержим переводы
                    current_trans = dict(rec.translations or {})
                    current_trans.update(data['translations'])
                    rec.translations = current_trans
                    
                    # Обратная совместимость колонок
                    rec.ru = data['translations'].get('ru', rec.ru)
                    rec.it = data['translations'].get('it', rec.it)
                    rec.de = data['translations'].get('de', rec.de)
                    
                    db_updates_count += 1

            if db_updates_count > 0:
                await db.commit()

            return {
                "success": True,
                "message": f"Синхронизация завершена. Обновлено в БД: {db_updates_count}, обновлено в файле: {file_updates_count}.",
                "updated_db": db_updates_count,
                "updated_file": file_updates_count
            }

        except PermissionError:
            return {"success": False, "message": "Ошибка доступа: закройте файл Excel в других программах и попробуйте снова."}
        except Exception as e:
            await db.rollback()
            logger.error(f"Sync conflict error: {e}", exc_info=True)
            return {"success": False, "message": f"Ошибка синхронизации: {str(e)}"}

    async def get_runtime_context(self, force_update: bool = False) -> Dict[str, Any]:
        """
        Получает текущий контекст (слова, винк) из кэша БД или генерирует новый.
        Гарантирует атомарность обновления через asyncio.Lock.
        """
        async with _get_state_lock():
            # 1. Получаем настройки из БД через сервис
            interval = int(await get_setting(self.db, 'max_random_minutes', '60'))
            last_update_str = await get_setting(self.db, 'last_update_ts', None)

            # 2. Проверка времени
            need_update = force_update
            now = datetime.now(timezone.utc)

            if not need_update:
                # Пытаемся взять текущий кэш из БД (если есть)
                current_words_json = await get_setting(self.db, 'current_words_cache', None)
                if not last_update_str or not current_words_json:
                    need_update = True
                else:
                    try:
                        last_dt = datetime.fromisoformat(last_update_str)
                        if last_dt.tzinfo is None: last_dt = last_dt.replace(tzinfo=timezone.utc)
                        diff = (now - last_dt).total_seconds() / 60
                        if diff >= interval:
                            need_update = True
                    except Exception as e:
                        logger.warning(f"Time parsing failed: {e}")
                        need_update = True

            # 3. Если время не пришло — отдаем из "кэша" в БД
            if not need_update:
                try:
                    words_str = await get_setting(self.db, 'current_words_cache', '[]')
                    words = json.loads(words_str if words_str else '[]')
                    wink = await get_setting(self.db, 'current_wink_cache', '...')
                    count = int(await get_setting(self.db, 'total_words_count', '0'))
                    imw = float(await get_setting(self.db, 'current_imw_cache', '0'))
                    coverage = float(await get_setting(self.db, 'current_coverage_cache', '100.0'))
                    return {'words': words, 'wink': wink, 'count': count, 'coverage': coverage, 'imw': imw}
                except Exception as e:
                    logger.warning(f"Cache read failed: {e}")
                    need_update = True

            # 4. Время пришло - Генерируем новые данные!
            return await self._update_runtime_context_from_db()

    async def _update_runtime_context_from_db(self) -> Dict[str, Any]:
        """Непосредственно генерирует новые данные из БД и сохраняет в кэш."""
        db = self.db
        now = datetime.now(timezone.utc)
        final_words: List[Dict[str, Any]] = []
        coverage, imw, total_count = 0.0, 0.0, 0
        try:
            from ..services.word_service import WordService
            word_service = WordService(db)
            metrics = await word_service.get_current_metrics()
            
            total_count = metrics['total_count']
            coverage = metrics['coverage']
            imw = metrics['imw']

            # Выбираем 3 случайных слова из всей базы
            active_langs = await word_service.get_active_languages()
            
            # Формируем условие: слово показывается, если ХОТЯ БЫ ОДИН активный язык НЕ известен
            lang_conditions = [
                text(f"JSON_EXTRACT(knowledge_stats, '$.{lang}') IS NOT 1") 
                for lang in active_langs
            ]
            
            stmt = select(models.WordStats).where(
                or_(*lang_conditions),
                models.WordStats.is_learned == False
            ).order_by(func.random()).limit(3)
            
            selected_res = await db.execute(stmt)
            selected = selected_res.scalars().all()
            
            # Получаем активные языки для кэширования правильной структуры
            active_langs_raw = await get_setting(db, 'active_languages', 'en,it,de')
            active_langs_list = [l.strip() for l in (active_langs_raw or 'en,it,de').split(',') if l.strip()]
            
            final_words = []
            for w in selected:
                word_data = {
                    'eng': w.eng, 
                    'ru': w.ru, 
                    'meaning': w.meaning, 
                    'is_learned': w.is_learned,
                    'translations': w.translations or {}
                }
                # Добавляем конкретные поля для обратной совместимости шаблонов (если они еще не обновлены)
                word_data['it'] = w.it
                word_data['de'] = w.de
                final_words.append(word_data)

            for w in selected:
                w.count += 1
                w.last_shown = now.replace(tzinfo=None) # SQLite stores naive
                
                # Обновляем поэзыковые счетчики в JSON
                stats = dict(w.show_stats or {})
                # Инкрементируем для всех активных иностранных языков + базовые
                for lang in active_langs_list:
                    stats[lang] = stats.get(lang, 0) + 1
                if 'en' not in stats: stats['en'] = stats.get('en', 0) + 1
                if 'ru' not in stats: stats['ru'] = stats.get('ru', 0) + 1
                w.show_stats = stats

            # Записываем ежедневную статистику показов
            today_date = now.date()
            daily_res = await db.execute(
                select(models.WordShowsDaily).where(models.WordShowsDaily.date == today_date)
            )
            daily_row = daily_res.scalar_one_or_none()
            if daily_row:
                daily_row.shows_count += len(selected)
            else:
                db.add(models.WordShowsDaily(date=today_date, shows_count=len(selected)))

            await db.commit()
        except Exception as e:
            logger.error(f"Metrics/Word Selection failed: {e}")
            await db.rollback()

        try:
            # Выбираем один случайный Wink прямо через БД
            wink_res = await db.execute(select(models.Wink).order_by(func.random()).limit(1))
            wink_obj = wink_res.scalar_one_or_none()
            wink_title = wink_obj.title if wink_obj else "..."
        except Exception as e:
            logger.warning(f"Wink selection failed: {e}")
            wink_title = "..."

        # 5. Сохраняем в кэш БД
        await set_settings_batch(db, {
            'current_words_cache': json.dumps(final_words, ensure_ascii=False),
            'current_wink_cache': wink_title,
            'total_words_count': str(total_count),
            'current_coverage_cache': str(coverage),
            'current_imw_cache': str(imw),
            'last_update_ts': now.isoformat()
        })

        return {'words': final_words, 'wink': wink_title, 'count': total_count, 'coverage': coverage, 'imw': imw}
